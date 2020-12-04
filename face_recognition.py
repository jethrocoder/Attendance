# Recognise Faces using some classification algorithm - like Logistic, KNN, SVM etc.


# 1. load the training data (numpy arrays of all the persons)
		# x- values are stored in the numpy arrays
		# y-values we need to assign for each person
# 2. Read a video stream using opencv
# 3. extract faces out of it
# 4. use knn to find the prediction of face (int)
# 5. map the predicted id to name of the user 
# 6. Display the predictions on the screen - bounding box and name

import cv2
import numpy as np 
import os 
import openpyxl
import datetime
import pandas as pd
from tkinter import messagebox

########## KNN CODE ############
def distance(v1, v2):
	# Eucledian 
	return np.sqrt(((v1-v2)**2).sum())

def knn(train, test, k=3):
	dist = []
	
	for i in range(train.shape[0]):
		# Get the vector and label
		ix = train[i, :-1]
		iy = train[i, -1]
		# Compute the distance from test point
		d = distance(test, ix)
		dist.append([d, iy])
	# Sort based on distance and get top k
	dk = sorted(dist, key=lambda x: x[0])[:k]
	# Retrieve only the labels
	labels = np.array(dk)[:, -1]
	
	# Get frequencies of each label
	output = np.unique(labels, return_counts=True)
	# Find max frequency and corresponding label
	index = np.argmax(output[1])
	return output[0][index]
################################

def mark_attendance():

	#Init Camera
	cap = cv2.VideoCapture(0)

	# Face Detection
	face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_alt.xml")

	skip = 0
	dataset_path = './data/'

	face_data = [] 
	labels = []     # we will use roll number as labels

	# class_id = 0 # Labels for the given file
	names = {} #Mapping btw rollno - name



	# Data Preparation

	student_table=pd.read_csv('student_details.csv')
	for fx in os.listdir(dataset_path):
		if fx.endswith('.npy'):
			#Create a mapping btw class_id and name
			rollNo = fx[:-4]
			name=student_table[student_table['roll']==rollNo].iloc[0]['name']
			names[rollNo] = name
			print('loaded '+name)
			data_item = np.load(dataset_path+fx)
			face_data.append(data_item)

			#Create Labels for the class
			target = np.array([rollNo for _ in range(data_item.shape[0])],dtype=object)

			labels.append(target)

	face_dataset = np.concatenate(face_data,axis=0)
	face_labels = np.concatenate(labels,axis=0).reshape((-1,1))

	print(face_dataset.shape)
	print(face_labels.shape)

	trainset = np.concatenate((face_dataset,face_labels),axis=1)
	print(trainset.shape)

	# Testing 

	pred_name=""
	while True:
		ret,frame = cap.read()
		if ret == False:
			continue

		frame = cv2.cvtColor(frame,cv2.COLOR_BGR2GRAY)
		frame = cv2.rotate(frame,cv2.ROTATE_180)

		faces = face_cascade.detectMultiScale(frame,1.3,5)
		if len(faces)==0:
			# print('your face is not visible \n please get into the frame')
			continue

		for face in faces:
			x,y,w,h = face

			#Get the face ROI
			offset = 10
			face_section = frame[y-offset:y+h+offset,x-offset:x+w+offset]
			face_section = cv2.resize(face_section,(100,100))

			#Predicted Label (out)
			out = knn(trainset,face_section.flatten())

			#Display on the screen the name and rectangle around it
			pred_name = names.get(out)
			cv2.putText(frame,pred_name,(x,y-10),cv2.FONT_HERSHEY_SIMPLEX,1,(255,0,0),2,cv2.LINE_AA)
			cv2.rectangle(frame,(x,y),(x+w,y+h),(0,255,255),2)

		cv2.imshow("Faces",frame)

		key = cv2.waitKey(1) & 0xFF
		if key==ord('q'):
			break

	# Marking attendance of student in attendance sheet
	cap.release()
	cv2.destroyAllWindows()

	def load_workbook(wb_path):
		if os.path.exists(wb_path):
			return openpyxl.load_workbook(wb_path)
		return openpyxl.Workbook()

	# student_table=pd.read_csv('student_details.csv')
	is_registered=True
	# for index, row in student_table.iterrows(): 
	# 	if row['Name']==pred_name:
	# 		is_registered=True
	# 		break

	# if is_registered==False:
	# 	cap.release()
	# 	cv2.destroyAllWindows()
	# 	messagebox.showinfo("Notification", "Student data not available\nPlease Register yourself!!") 
	# 	return
		
	roll=student_table[student_table['name']==pred_name].iloc[0]['roll']
	dt=datetime.datetime.now()
	date = dt.strftime("%m-%d-%Y")
	time = dt.strftime("%H:%M:%S")
	filename=date
	wb_path='Attendance/'+filename+'.xlsx'
	print(wb_path)
	wb=load_workbook(wb_path)
	sheet=wb['Sheet']
	sheet.append((pred_name,roll,time))
	wb.save(wb_path)

	

	messagebox.showinfo("Notification", "Your attendance is marked") 
